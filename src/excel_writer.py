import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


FIELDNAMES = [
    "company_name",
    "screener_url",
    "result_pdf_link",
    "price",
    "market_cap_text",
    "market_cap_cr",
    "pe",
    "sales_latest_qtr_cr",
    "yoy_revenue_pct",
    "net_profit_latest_qtr_cr",
    "yoy_profit_pct",
    "pb",
    "qoq_revenue_pct",
    "qoq_profit_pct",
    "debt_to_equity",
    "roce",
    "roe",
    "comp_sales_growth_10y",
    "comp_sales_growth_5y",
    "comp_sales_growth_3y",
    "comp_sales_growth_ttm",
    "comp_profit_growth_10y",
    "comp_profit_growth_5y",
    "comp_profit_growth_3y",
    "comp_profit_growth_ttm",
    "stock_price_cagr_10y",
    "stock_price_cagr_5y",
    "stock_price_cagr_3y",
    "stock_price_cagr_1y",
    "return_on_equity_10y",
    "return_on_equity_5y",
    "return_on_equity_3y",
    "return_on_equity_last_year",
    "promoter_latest",
    "promoter_1y_ago",
    "promoter_change_1y",
    "fii_latest",
    "fii_1y_ago",
    "fii_change_1y",
    "shareholding_trend_note",
]

HEADERS = [
    "Company Name",
    "Screener URL",
    "PDF URL",
    "Price",
    "Market Cap",
    "Market Cap (Cr)",
    "PE",
    "Sales Latest Qtr (Cr)",
    "Sales YoY (%)",
    "Net Profit Latest Qtr (Cr)",
    "Net Profit YoY (%)",
]


def write_csv(rows, path):
    if not rows:
        return

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_excel(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered Results"

    ws.append(HEADERS)

    for row in rows:
        ws.append([
            row.get("company_name"),
            row.get("screener_url"),
            row.get("result_pdf_link"),
            row.get("price"),
            row.get("market_cap_text"),
            row.get("market_cap_cr"),
            row.get("pe"),
            row.get("sales_latest_qtr_cr"),
            row.get("yoy_revenue_pct"),
            row.get("net_profit_latest_qtr_cr"),
            row.get("yoy_profit_pct"),
        ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri")
    body_font = Font(name="Calibri")
    link_font = Font(name="Calibri", color="0000EE", underline="single")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.alignment = left if col_idx in (1, 2, 3, 5) else right

        for col_idx in (2, 3):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = link_font
                cell.alignment = left

        for col_idx in (4, 6, 7, 8, 9, 10, 11):
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'

    widths = {
        "A": 28, "B": 42, "C": 42, "D": 12, "E": 16,
        "F": 16, "G": 10, "H": 20, "I": 14, "J": 23, "K": 18
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"

    if ws.max_row > 1:
        table = Table(displayName="FilteredResults", ref=f"A1:K{ws.max_row}")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    wb.save(path)