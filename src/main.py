from __future__ import annotations

from pathlib import Path
from datetime import datetime, date
from urllib.parse import urlencode, urljoin
import csv
import os
import re

import requests
from bs4 import BeautifulSoup

from filters import parse_market_cap_to_cr, market_cap_above

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False


SCREENER_BASE = "https://www.screener.in"
LOGIN_URL = f"{SCREENER_BASE}/login/"
LATEST_RESULTS_URL = f"{SCREENER_BASE}/results/latest/"
MIN_MARKET_CAP_CR = 500


BASE = Path.cwd()
META_DIR = BASE / "data" / "metadata"
META_DIR.mkdir(parents=True, exist_ok=True)

DEBUG_FILE = META_DIR / "debug.txt"


def log(msg: str):
    print(msg)
    with DEBUG_FILE.open("a", encoding="utf-8") as f:
        f.write(msg + "\n")


def write_run_log(status: str, notes: str):
    file_path = META_DIR / "runs.csv"
    exists = file_path.exists()
    with file_path.open("a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if not exists:
            writer.writerow(["run_utc", "status", "notes"])
        writer.writerow([datetime.utcnow().isoformat() + "Z", status, notes])


def save_text(filename: str, content: str):
    out = META_DIR / filename
    out.write_text(content, encoding="utf-8")
    return out


def write_results_csv(rows, filename: str):
    out = META_DIR / filename
    with out.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "company_name",
                "screener_url",
                "market_cap_text",
                "market_cap_cr",
                "result_pdf_link",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)
    return out


def write_results_xlsx(rows, filename: str):
    if not OPENPYXL_AVAILABLE:
        log("openpyxl not available, skipping xlsx generation")
        return None

    out = META_DIR / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Calibri", size=11, color="000000")
    link_font = Font(name="Calibri", size=11, color="0000FF", underline="single")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    headers = [
        "Company Name",
        "Screener URL",
        "Market Cap",
        "Market Cap (Cr)",
        "Result PDF Link",
    ]
    ws.append(headers)

    for row in rows:
        ws.append([
            row.get("company_name", ""),
            row.get("screener_url", ""),
            row.get("market_cap_text", ""),
            row.get("market_cap_cr", ""),
            row.get("result_pdf_link", ""),
        ])

    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center

    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).font = body_font
        ws.cell(r, 1).alignment = left

        c2 = ws.cell(r, 2)
        c2.font = link_font
        c2.alignment = left
        if c2.value:
            c2.hyperlink = c2.value

        c3 = ws.cell(r, 3)
        c3.font = body_font
        c3.alignment = right

        c4 = ws.cell(r, 4)
        c4.font = body_font
        c4.alignment = right
        c4.number_format = "#,##0.00"

        c5 = ws.cell(r, 5)
        c5.font = link_font
        c5.alignment = left
        if c5.value:
            c5.hyperlink = c5.value

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 45

    ws.freeze_panes = "A2"

    if ws.max_row >= 2:
        table = Table(displayName="ScreenerResults", ref=f"A1:E{ws.max_row}")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    wb.save(out)
    return out


def get_csrf_token(html: str) -> str:
    soup = BeautifulSoup(html, "lxml")
    el = soup.find("input", {"name": "csrfmiddlewaretoken"})
    return el.get("value", "") if el else ""


def login(session: requests.Session, email: str, password: str):
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": LOGIN_URL,
    }

    r = session.get(LOGIN_URL, headers=headers, timeout=30)
    r.raise_for_status()
    save_text("screener_login_page.html", r.text)

    csrf_token = get_csrf_token(r.text)
    if not csrf_token:
        raise RuntimeError("Could not find csrfmiddlewaretoken on login page")

    payload = {
        "csrfmiddlewaretoken": csrf_token,
        "username": email,
        "password": password,
        "next": "",
    }

    post_headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": LOGIN_URL,
        "Origin": SCREENER_BASE,
    }

    resp = session.post(
        LOGIN_URL,
        data=payload,
        headers=post_headers,
        timeout=30,
        allow_redirects=True,
    )
    resp.raise_for_status()
    save_text("screener_login_response.html", resp.text)

    low = resp.text.lower()
    if "login to your account" in low or "get a free account" in low:
        raise RuntimeError("Login failed; still seeing login page")


def fetch_html(url: str, session: requests.Session) -> str:
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": SCREENER_BASE,
    }
    r = session.get(url, headers=headers, timeout=30)
    r.raise_for_status()
    return r.text


def build_results_url(d: date | None = None) -> str:
    params = {"all": ""}

    if d:
        params["result_update_date__day"] = d.day
        params["result_update_date__month"] = d.month
        params["result_update_date__year"] = d.year

    qs = urlencode(params)
    return f"{LATEST_RESULTS_URL}?{qs}"


def normalize_company_url(href: str) -> str:
    return urljoin(SCREENER_BASE, href.split("?")[0])


def normalize_url(href: str) -> str:
    return urljoin(SCREENER_BASE, href)


def looks_like_company_href(href: str) -> bool:
    if not href:
        return False
    if not href.startswith("/company/"):
        return False
    if "/source/quarter/" in href:
        return False
    return True


def clean_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()


def extract_market_cap_from_text(text: str):
    text = clean_text(text)
    if not text:
        return "", None

    m = re.search(r"(?:m\.?\s*cap|market\s*cap)\s*[:\-]?\s*([\d,]+(?:\.\d+)?)\s*(Cr|Crore|Lac|Lakh|Bn|Billion)?", text, re.I)
    if m:
        raw_num = m.group(1)
        raw_unit = m.group(2) or "Cr"
        market_cap_text = f"{raw_num} {raw_unit}"
        market_cap_cr = parse_market_cap_to_cr(market_cap_text)
        return market_cap_text, market_cap_cr

    m2 = re.search(r"([\d,]+(?:\.\d+)?)\s*(Cr|Crore|Lac|Lakh|Bn|Billion)", text, re.I)
    if m2:
        market_cap_text = f"{m2.group(1)} {m2.group(2)}"
        market_cap_cr = parse_market_cap_to_cr(market_cap_text)
        return market_cap_text, market_cap_cr

    return "", None


def extract_rows_from_latest(html: str):
    soup = BeautifulSoup(html, "lxml")
    rows = []
    seen = set()

    trs = soup.find_all("tr")
    log(f"Found {len(trs)} <tr> rows")

    for tr in trs:
        anchors = tr.find_all("a", href=True)

        company_a = None
        pdf_a = None

        for a in anchors:
            href = a.get("href", "").strip()
            if looks_like_company_href(href) and not company_a:
                company_a = a
            if "/company/source/quarter/" in href and not pdf_a:
                pdf_a = a

        if not company_a:
            continue

        company_name = clean_text(company_a.get_text(" ", strip=True))
        company_url = normalize_company_url(company_a.get("href", "").strip())

        if not company_name or company_name.upper() == "PDF":
            continue

        if company_url in seen:
            continue
        seen.add(company_url)

        row_text = clean_text(tr.get_text(" ", strip=True))
        market_cap_text, market_cap_cr = extract_market_cap_from_text(row_text)

        result_pdf_link = ""
        if pdf_a:
            result_pdf_link = normalize_url(pdf_a.get("href", "").strip())

        row = {
            "company_name": company_name,
            "screener_url": company_url,
            "market_cap_text": market_cap_text,
            "market_cap_cr": market_cap_cr,
            "result_pdf_link": result_pdf_link,
        }
        rows.append(row)

    return rows


def main():
    DEBUG_FILE.write_text("", encoding="utf-8")

    log(f"CWD = {Path.cwd()}")
    log(f"main.py = {Path(__file__).resolve()}")
    log(f"META_DIR = {META_DIR}")
    log(f"META_DIR exists = {META_DIR.exists()}")

    (META_DIR / "debug.txt").write_text("main.py started\n", encoding="utf-8")

    email = os.getenv("SCREENER_EMAIL", "").strip()
    password = os.getenv("SCREENER_PASSWORD", "").strip()
    date_str = os.getenv("SCREENER_RESULTS_DATE", "").strip()

    if not email or not password:
        write_run_log("error", "Missing SCREENER_EMAIL / SCREENER_PASSWORD")
        raise RuntimeError("Missing SCREENER_EMAIL / SCREENER_PASSWORD")

    d = date.fromisoformat(date_str) if date_str else None
    mode = f"date={d.isoformat()}" if d else "date=all"
    log(f"Mode = {mode}")

    session = requests.Session()
    login(session, email, password)
    log("Login successful")

    url = build_results_url(d=d)
    log(f"Results URL = {url}")

    html = fetch_html(url, session)
    save_text("screener_latest_results.html", html)
    log(f"Saved HTML, length = {len(html)}")

    rows = extract_rows_from_latest(html)
    log(f"Extracted rows = {len(rows)}")

    if rows:
        log(f"Sample row = {rows[0]}")

    all_csv = write_results_csv(rows, "screener_results_all.csv")
    log(f"Wrote CSV = {all_csv}")

    filtered_rows = market_cap_above(rows, MIN_MARKET_CAP_CR)
    log(f"Filtered rows (market_cap > {MIN_MARKET_CAP_CR} Cr) = {len(filtered_rows)}")

    filtered_csv = write_results_csv(filtered_rows, "screener_results_filtered.csv")
    log(f"Wrote filtered CSV = {filtered_csv}")

    xlsx_file = write_results_xlsx(filtered_rows, "screener_results_filtered.xlsx")
    if xlsx_file:
        log(f"Wrote XLSX = {xlsx_file}")

    write_run_log(
        "ok",
        f"screener_logged_in mode={mode} total_rows={len(rows)} filtered_rows={len(filtered_rows)} min_market_cap_cr={MIN_MARKET_CAP_CR}"
    )

    log("Final files in META_DIR:")
    for p in sorted(META_DIR.glob("*")):
        log(f"- {p.name}")


if __name__ == "__main__":
    main()